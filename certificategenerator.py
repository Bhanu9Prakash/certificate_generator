import cv2 as cv
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


mail_content = '''Hello,
Thank you for participating in XYZ event.
Hope to see you again next year with the same Enthusiasm.
Please find your digital certificate attached.
Thank You
'''


template_path = '/Users/yathartharora/certificate_template/templates/template1.png'
details_path = '/Users/yathartharora/certificate_template/details.xlsx'
output_path = '/Users/yathartharora/certificate_template/generated_certificates/'


sender_address = 'concerncedauthority@gmail.com'
sender_pass = 'Blackred@123'


font_size = 3
font_color = (0,0,0)


coordinate_y_adjustment = 15
coordinate_x_adjustment = 7


obj = openpyxl.load_workbook(details_path)
sheet = obj.active


for i in range(3,5):

    get_name = sheet.cell(row = i ,column = 1)
    certi_name = get_name.value
    get_email = sheet.cell(row = i ,column = 2)
    receiver_address = get_email.value
    
    
    img = cv.imread(template_path)

    font = cv.FONT_HERSHEY_PLAIN

    text_size = cv.getTextSize(certi_name, font, font_size, 10)[0]
    text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
    text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)



    cv.putText(img, certi_name, (text_x ,text_y ), font, font_size, font_color, 10)

    certi_path = output_path + certi_name + '.png'

    cv.imwrite(certi_path,img)
    
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = 'Digital Certificate'
    message.attach(MIMEText(mail_content, 'plain'))
    attach_file_name = output_path + certi_name + '.png'
    print(attach_file_name)
    attach_file = open(attach_file_name, 'rb')
    payload = MIMEBase('application', 'octate-stream')
    payload.set_payload((attach_file).read())
    encoders.encode_base64(payload)
    payload.add_header('Content-Decomposition', 'attachment', filename=attach_file_name)
    message.attach(payload)
    
    session = smtplib.SMTP('smtp.gmail.com', 587)
    session.starttls()
    session.login(sender_address, sender_pass)
    session.sendmail(sender_address, receiver_address, message.as_string())
    session.quit()
    print('Mail Sent')

cv.destroyAllWindows()
